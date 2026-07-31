"""
관세청 HS부호 단위별 품목명 엑셀(HS2/HS4/HS6 단위 시트)에서 특정 류(chapter)의
품목명을 뽑아 products_real.csv와 동일한 스키마(description, hs_code,
confidence_basis)로 변환하는 스크립트.

실험 3(docs/09-experiment3-plan.md)의 데이터 다양화용 — 61/62류(의류)에 편중된
기존 products_real.csv에 63/64/84/85류 등 다른 류를 추가하기 위해 사용한다.

사용 예:
  python src/finetune/extract_hs_reference.py \
      --xlsx "C:/Users/user01/Desktop/관세청_HS부호 단위별 품목명_20260101.xlsx" \
      --chapters 63,64,84,85 --per-chapter 40 \
      --output data/raw/products_new_chapters.csv
"""

import argparse
import csv
import random
from pathlib import Path

import openpyxl

HS2_SHEET = "HS2단위"
HS4_SHEET = "HS4단위"
HS6_SHEET = "HS6단위(5단위포함)"


def load_hs2(ws) -> dict[str, str]:
    out = {}
    for code, kor, _eng in ws.iter_rows(min_row=2, values_only=True):
        out[str(code)] = kor.strip()
    return out


def load_hs4(ws) -> dict[str, tuple[str, str]]:
    out = {}
    for code, kor, eng in ws.iter_rows(min_row=2, values_only=True):
        out[str(code)] = (kor.strip(), eng.strip())
    return out


def load_hs6_entries(ws) -> list[tuple[str, str, str]]:
    """(code, kor, eng) 리스트. code는 5자리(중간 그룹) or 6자리(최종 소호) 모두 포함."""
    out = []
    for code, kor, eng in ws.iter_rows(min_row=2, values_only=True):
        code = str(code)
        if kor is None or eng is None:
            continue
        out.append((code, kor.strip(), eng.strip()))
    return out


def short_chapter_name(hs2_kor: str) -> str:
    # "제61류 편물제(編物製)나 뜨개질 편물제인 의류와 그 부속품" -> "편물제(編物製)나 뜨개질 편물제인 의류와 그 부속품"
    if "류 " in hs2_kor:
        name = hs2_kor.split("류 ", 1)[1]
    else:
        name = hs2_kor
    return name[:15] + ("..." if len(name) > 15 else "")


def truncated(text: str, limit: int = 20) -> str:
    return text[:limit] + ("..." if len(text) > limit else "")


def build_rows(
    chapters: list[str],
    per_chapter: int,
    hs2_map: dict[str, str],
    hs4_map: dict[str, tuple[str, str]],
    hs6_entries: list[tuple[str, str, str]],
    seed: int,
) -> list[dict]:
    rng = random.Random(seed)
    rows = []
    for ch in chapters:
        candidates = [e for e in hs6_entries if e[0][:2] == ch]
        rng.shuffle(candidates)
        selected = candidates[:per_chapter]

        hs2_kor = hs2_map.get(ch, f"제{ch}류")
        chapter_short = short_chapter_name(hs2_kor)

        for code, kor, eng in selected:
            heading = code[:4]
            if heading not in hs4_map:
                continue
            heading_kor, heading_eng = hs4_map[heading]

            eng_clean = eng.rstrip(" :").rstrip(".")
            heading_eng_clean = heading_eng.rstrip(" :").rstrip(".")
            description = f"{heading_eng_clean} - {eng_clean}"

            kor_clean = kor.rstrip(" :")
            confidence_basis = (
                f"제{ch}류({chapter_short}) > "
                f"제{heading}호({truncated(heading_kor)}) > "
                f"{code}({kor_clean})"
            )

            rows.append(
                {
                    "description": description,
                    "hs_code": code,
                    "confidence_basis": confidence_basis,
                }
            )
    return rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--chapters", type=str, required=True, help="쉼표로 구분된 2자리 류 코드, 예: 63,64,84,85")
    parser.add_argument("--per-chapter", type=int, default=40, help="류당 최대 샘플 수")
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--output", type=Path, default=Path("data/raw/products_new_chapters.csv"))
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    chapters = [c.strip() for c in args.chapters.split(",") if c.strip()]

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    hs2_map = load_hs2(wb[HS2_SHEET])
    hs4_map = load_hs4(wb[HS4_SHEET])
    hs6_entries = load_hs6_entries(wb[HS6_SHEET])

    rows = build_rows(chapters, args.per_chapter, hs2_map, hs4_map, hs6_entries, args.seed)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["description", "hs_code", "confidence_basis"])
        writer.writeheader()
        writer.writerows(rows)

    per_ch_count = {ch: sum(1 for r in rows if r["hs_code"][:2] == ch) for ch in chapters}
    print(f"총 {len(rows)}건 저장: {args.output}")
    print("류별 건수:", per_ch_count)
